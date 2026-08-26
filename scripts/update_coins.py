from pathlib import Path
import json
import re
import shutil

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent

EXCEL_FILE = ROOT / "data" / "Euro_Coin_Catalog_Automation.xlsx"
JSON_FILE = ROOT / "data" / "coins.json"

INCOMING_DIR = ROOT / "incoming"
IMAGES_DIR = ROOT / "images" / "coins"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def number_or_zero(value):
    if value in (None, ""):
        return 0

    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def year_value(value):
    if value in (None, ""):
        return None

    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def safe_filename_part(value):
    value = clean(value)

    value = value.replace("€", "Euro")
    value = value.replace("&", "and")
    value = value.replace("/", "-")
    value = value.replace("\\", "-")
    value = value.replace("'", "")
    value = value.replace('"', "")

    value = re.sub(r"\s+", "-", value)
    value = re.sub(r"-+", "-", value)
    value = re.sub(r"[^A-Za-z0-9À-ž_-]", "", value)

    return value.strip("-_")


def denomination_filename(value):
    value = clean(value)
    value = value.replace("€", "Euro")
    value = value.replace(" ", "")

    return safe_filename_part(value)


def build_image_filename(
    coin_id,
    country,
    year,
    denomination,
    coin_type,
    name
):
    parts = [
        coin_id,
        safe_filename_part(country),
        str(year),
        denomination_filename(denomination),
        safe_filename_part(coin_type or "Unknown")
    ]

    if coin_type == "Commemorative" and name:
        parts.append(
            safe_filename_part(name)
        )

    return "_".join(parts) + ".jpeg"


def load_country_codes(workbook):
    sheet = workbook["Lists"]

    mapping = {}

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):
        country = clean(row[0])
        code = clean(row[1])

        if country and code:
            mapping[country] = code

    return mapping


def load_specifications(workbook):
    sheet = workbook["Specifications"]

    specs = {}

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):
        denomination = clean(row[0])

        if not denomination:
            continue

        specs[denomination] = {
            "metal": clean(row[1]),
            "weight": clean(row[2]),
            "diameter": clean(row[3])
        }

    return specs


def get_headers(sheet):
    headers = {}

    for column, cell in enumerate(
        sheet[1],
        start=1
    ):
        name = clean(cell.value)

        if name:
            headers[name] = column

    return headers


def get_value(
    sheet,
    row,
    headers,
    name
):
    column = headers.get(name)

    if not column:
        return ""

    return sheet.cell(
        row=row,
        column=column
    ).value


def find_existing_image(
    filename,
    country,
    year,
    denomination,
    coin_type
):
    country_folder = (
        IMAGES_DIR /
        safe_filename_part(country)
    )

    destination = (
        country_folder /
        filename
    )

    if destination.exists():
        return destination

    if not country_folder.exists():
        return None

    suffix = (
        f"_{safe_filename_part(country)}"
        f"_{year}"
        f"_{denomination_filename(denomination)}"
        f"_{safe_filename_part(coin_type)}"
    )

    for file in country_folder.iterdir():
        if not file.is_file():
            continue

        if (
            file.suffix.lower() in
            {".jpeg", ".jpg", ".png", ".webp"}
            and suffix in file.stem
        ):
            return file

    return None


def process_image(
    filename,
    country,
    year,
    denomination,
    coin_type
):
    country_folder = (
        IMAGES_DIR /
        safe_filename_part(country)
    )

    country_folder.mkdir(
        parents=True,
        exist_ok=True
    )

    destination = (
        country_folder /
        filename
    )

    source = (
        INCOMING_DIR /
        filename
    )

    if source.exists():

        if destination.exists():
            destination.unlink()

        shutil.move(
            str(source),
            str(destination)
        )

        print(
            "Moved image: "
            f"{source.relative_to(ROOT)} "
            "-> "
            f"{destination.relative_to(ROOT)}"
        )

        return destination.relative_to(
            ROOT
        ).as_posix()

    existing = find_existing_image(
        filename,
        country,
        year,
        denomination,
        coin_type
    )

    if existing:
        return existing.relative_to(
            ROOT
        ).as_posix()

    print(
        "WARNING: image not found: "
        f"{filename}"
    )

    return ""


def main():

    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Excel file not found: "
            f"{EXCEL_FILE}"
        )

    INCOMING_DIR.mkdir(
        parents=True,
        exist_ok=True
    )

    IMAGES_DIR.mkdir(
        parents=True,
        exist_ok=True
    )

    workbook = load_workbook(
        EXCEL_FILE,
        data_only=True
    )

    coins_sheet = workbook["Coins"]

    country_codes = (
        load_country_codes(workbook)
    )

    specifications = (
        load_specifications(workbook)
    )

    headers = get_headers(
        coins_sheet
    )

    required_columns = [
        "Country",
        "Year",
        "Denomination",
    ]

    missing_columns = [
        column
        for column in required_columns
        if column not in headers
    ]

    if missing_columns:
        raise RuntimeError(
            "Missing Excel columns: "
            + ", ".join(missing_columns)
        )

    coins = []

    for row in range(
        2,
        coins_sheet.max_row + 1
    ):

        country = clean(
            get_value(
                coins_sheet,
                row,
                headers,
                "Country"
            )
        )

        # Completely blank row -> ignore
        if not country:
            continue

        year = year_value(
            get_value(
                coins_sheet,
                row,
                headers,
                "Year"
            )
        )

        denomination = clean(
            get_value(
                coins_sheet,
                row,
                headers,
                "Denomination"
            )
        )

        coin_type = clean(
            get_value(
                coins_sheet,
                row,
                headers,
                "Type"
            )
        )

        name = clean(
            get_value(
                coins_sheet,
                row,
                headers,
                "Name"
            )
        )

        condition = clean(
            get_value(
                coins_sheet,
                row,
                headers,
                "Condition"
            )
        )

        status = clean(
            get_value(
                coins_sheet,
                row,
                headers,
                "Status"
            )
        )

        duplicates = number_or_zero(
            get_value(
                coins_sheet,
                row,
                headers,
                "Duplicates"
            )
        )

        mintage = clean(
            get_value(
                coins_sheet,
                row,
                headers,
                "Mintage"
            )
        )

        description = clean(
            get_value(
                coins_sheet,
                row,
                headers,
                "Description"
            )
        )

        notes = clean(
            get_value(
                coins_sheet,
                row,
                headers,
                "Notes"
            )
        )

        if year is None:
            print(
                f"WARNING: skipped row "
                f"{row}: invalid year"
            )
            continue

        if not denomination:
            print(
                f"WARNING: skipped row "
                f"{row}: no denomination"
            )
            continue

        # 1 cent – 1 € are always Regular.
        # For 2 € the Excel value must be
        # Regular or Commemorative.
        if denomination != "2 €":
            coin_type = "Regular"

   if coin_type and coin_type not in {
    "Regular",
    "Commemorative"
}:
    print(
        f"WARNING: row {row}: invalid Type '{coin_type}', leaving it blank"
    )
    coin_type = ""

        if coin_type not in {
            "Regular",
            "Commemorative"
        }:
            print(
                f"WARNING: skipped row "
                f"{row}: invalid Type "
                f"'{coin_type}'"
            )
            continue

        status_lower = status.lower()

        if status_lower not in {
    "collection",
    "missing",
    "duplicate"
}:
    status_lower = "collection"

        # Duplicate must NOT also appear
        # in Missing.
        in_collection = (
            status_lower
            in {"collection", "duplicate"}
        )

        if status_lower == "duplicate":

            if duplicates <= 0:
                duplicates = 1

        else:
            duplicates = 0

        wanted = False

        coin_id = f"{row - 1:03d}"

        country_code = (
            country_codes.get(
                country,
                ""
            )
        )

        if not country_code:
            print(
                "WARNING: no CountryCode "
                f"for {country}"
            )

        spec = specifications.get(
            denomination,
            {}
        )

        metal = spec.get(
            "metal",
            ""
        )

        weight = spec.get(
            "weight",
            ""
        )

        diameter = spec.get(
            "diameter",
            ""
        )

        if not description:

            if coin_type == "Regular":
                description = (
                    f"Regular {denomination} "
                    f"circulation coin of "
                    f"{country}."
                )

            elif name:
                description = (
                    f"{name}. "
                    f"Commemorative 2 euro "
                    f"coin of {country}."
                )

        image_filename = (
            build_image_filename(
                coin_id,
                country,
                year,
                denomination,
                coin_type,
                name
            )
        )

        image_path = process_image(
            image_filename,
            country,
            year,
            denomination,
            coin_type
        )

        if name:
            display_name = name

        elif coin_type == "Regular":

            display_name = (
                denomination
                .replace("€", "Euro")
            )

        else:
            display_name = denomination

        coin = {
            "country": country,
            "countryCode": country_code,
            "year": year,
            "denomination": denomination,
            "type": coin_type,
            "name": display_name,
            "condition": condition,
            "mint": "",
            "inCollection": in_collection,
            "duplicates": duplicates,
            "wanted": wanted,
            "image": image_path,
            "mintage": mintage,
            "metal": metal,
            "weight": weight,
            "diameter": diameter,
            "description": description,
            "notes": notes
        }

        coins.append(coin)

    JSON_FILE.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    with JSON_FILE.open(
        "w",
        encoding="utf-8"
    ) as file:

        json.dump(
            coins,
            file,
            ensure_ascii=False,
            indent=2
        )

        file.write("\n")

    print()
    print(
        "Generated: "
        f"{JSON_FILE.relative_to(ROOT)}"
    )
    print(
        f"Coins: {len(coins)}"
    )


if __name__ == "__main__":
    main()
