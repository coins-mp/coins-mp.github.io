from pathlib import Path
from copy import copy
import json
import re

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent

SOURCE_JSON_FILE = ROOT / "data" / "coins-source.json"
EXCEL_FILE = ROOT / "data" / "Euro_Coin_Catalog_Automation.xlsx"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def safe_filename_part(value):
    value = clean(value)

    value = value.replace("€", "Euro")
    value = value.replace("&", "and")
    value = value.replace("/", "-")
    value = value.replace("\\", "-")
    value = value.replace("'", "")
    value = value.replace('"', "")

    value = re.sub(r"\s+", "-", value)
    value = re.sub(r"-+", "-", value)
    value = re.sub(
        r"[^A-Za-z0-9À-ž_-]",
        "",
        value
    )

    return value.strip("-_")


def denomination_filename(value):
    value = clean(value)
    value = value.replace("€", "Euro")
    value = value.replace(" ", "")

    return safe_filename_part(value)


def build_image_filename(
    coin_id,
    country,
    year,
    denomination,
    coin_type,
    name
):
    parts = [
        coin_id,
        safe_filename_part(country),
        str(year),
        denomination_filename(denomination),
        safe_filename_part(
            coin_type or "Unknown"
        )
    ]

    if (
        coin_type == "Commemorative"
        and clean(name)
    ):
        parts.append(
            safe_filename_part(name)
        )

    return "_".join(parts) + ".jpeg"


def load_country_codes(workbook):
    sheet = workbook["Lists"]

    result = {}

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):
        country = clean(row[0])
        code = clean(row[1])

        if country and code:
            result[country] = code

    return result


def load_specifications(workbook):
    sheet = workbook["Specifications"]

    result = {}

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):
        denomination = clean(row[0])

        if not denomination:
            continue

        result[denomination] = {
            "metal": clean(row[1]),
            "weight": clean(row[2]),
            "diameter": clean(row[3])
        }

    return result


def copy_row_style(
    sheet,
    source_row,
    target_row,
    max_column
):
    for column in range(
        1,
        max_column + 1
    ):
        source = sheet.cell(
            row=source_row,
            column=column
        )

        target = sheet.cell(
            row=target_row,
            column=column
        )

        if source.has_style:
            target._style = copy(
                source._style
            )

        target.number_format = (
            source.number_format
        )

        target.font = copy(
            source.font
        )

        target.fill = copy(
            source.fill
        )

        target.border = copy(
            source.border
        )

        target.alignment = copy(
            source.alignment
        )

        target.protection = copy(
            source.protection
        )


def main():

    if not SOURCE_JSON_FILE.exists():
        raise FileNotFoundError(
            f"Source JSON not found: "
            f"{SOURCE_JSON_FILE}"
        )

    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Excel template not found: "
            f"{EXCEL_FILE}"
        )

    with SOURCE_JSON_FILE.open(
        "r",
        encoding="utf-8"
    ) as file:
        coins = json.load(file)

    workbook = load_workbook(
        EXCEL_FILE,
        data_only=False
    )

    sheet = workbook["Coins"]

    country_codes = (
        load_country_codes(workbook)
    )

    specifications = (
        load_specifications(workbook)
    )

    headers = {}

    for column, cell in enumerate(
        sheet[1],
        start=1
    ):
        header = clean(cell.value)

        if header:
            headers[header] = column

    required_headers = [
        "ID",
        "Country",
        "CountryCode",
        "Year",
        "Denomination",
        "Type",
        "Name",
        "Condition",
        "Status",
        "Duplicates",
        "ImageFile",
        "Mintage",
        "Metal",
        "Weight",
        "Diameter",
        "Description",
        "Notes"
    ]

    missing_headers = [
        header
        for header in required_headers
        if header not in headers
    ]

    if missing_headers:
        raise RuntimeError(
            "Missing Excel headers: "
            + ", ".join(missing_headers)
        )

    #
    # Clear old catalog values.
    # Styles / validation remain in the workbook.
    #
    for row in range(
        2,
        sheet.max_row + 1
    ):
        for column in range(
            1,
            len(required_headers) + 1
        ):
            sheet.cell(
                row=row,
                column=column
            ).value = None

    #
    # Write current primary JSON data.
    #
    for index, coin in enumerate(
        coins,
        start=1
    ):
        excel_row = index + 1

        if excel_row > 2:
            copy_row_style(
                sheet,
                2,
                excel_row,
                len(required_headers)
            )

        coin_id = clean(
            coin.get("id")
        )

        if not coin_id:
            coin_id = f"{index:03d}"

        country = clean(
            coin.get("country")
        )

        year = coin.get("year")

        denomination = clean(
            coin.get("denomination")
        )

        coin_type = clean(
            coin.get("type")
        )

        name = clean(
            coin.get("name")
        )

        condition = clean(
            coin.get("condition")
        )

        status = clean(
            coin.get("status")
        )

        mintage = clean(
            coin.get("mintage")
        )

        description = clean(
            coin.get("description")
        )

        notes = clean(
            coin.get("notes")
        )

        image_file = clean(
            coin.get("imageFile")
        )

        if not image_file:
            image_file = (
                build_image_filename(
                    coin_id,
                    country,
                    year,
                    denomination,
                    coin_type,
                    name
                )
            )

        country_code = (
            country_codes.get(
                country,
                ""
            )
        )

        specification = (
            specifications.get(
                denomination,
                {}
            )
        )

        metal = specification.get(
            "metal",
            ""
        )

        weight = specification.get(
            "weight",
            ""
        )

        diameter = specification.get(
            "diameter",
            ""
        )

        if not description:
            if coin_type == "Regular":
                description = (
                    f"Regular "
                    f"{denomination} "
                    f"circulation coin of "
                    f"{country}."
                )

            elif (
                coin_type
                == "Commemorative"
                and name
            ):
                description = (
                    f"{name}. "
                    f"Commemorative "
                    f"2 euro coin of "
                    f"{country}."
                )

        duplicate_value = ""

        if status.lower() == "duplicate":
            duplicate_value = 1

        values = {
            "ID": coin_id,
            "Country": country,
            "CountryCode": country_code,
            "Year": year,
            "Denomination": denomination,
            "Type": coin_type,
            "Name": name,
            "Condition": condition,
            "Status": status,
            "Duplicates": duplicate_value,
            "ImageFile": image_file,
            "Mintage": mintage,
            "Metal": metal,
            "Weight": weight,
            "Diameter": diameter,
            "Description": description,
            "Notes": notes
        }

        for header, value in values.items():
            sheet.cell(
                row=excel_row,
                column=headers[header]
            ).value = value

    #
    # Freeze header row.
    #
    sheet.freeze_panes = "A2"

    workbook.save(
        EXCEL_FILE
    )

    print(
        f"Excel generated from "
        f"coins-source.json"
    )

    print(
        f"Coins exported: "
        f"{len(coins)}"
    )

    print(
        f"Updated: "
        f"{EXCEL_FILE.relative_to(ROOT)}"
    )


if __name__ == "__main__":
    main()
